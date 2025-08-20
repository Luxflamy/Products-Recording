import streamlit as st
import pandas as pd
from datetime import datetime
import os
import csv
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

# ---------------------- 基础设置 ----------------------
st.set_page_config(page_title="退货登记与导出系统", layout="wide")

CSV_FILE = "退货.csv"
IMAGE_FOLDER = "uploaded_images"
DEFAULT_COLUMNS = {
    'timestamp': '退货时间',
    'tracking_number': '货件号',
    'product_name': '产品名称',
    'barcode': '条形码',
    'SAIN': 'SAIN码',
    'product_name_actual': '实际产品名称',
    'return_reason': '退货原因',
    'notes': '备注',
    'image_name': '图片文件名'
}
RETURN_REASONS = ["包装破损", "包装较脏", "运输损坏", "其他原因"]

def init_data():
    if not os.path.exists(CSV_FILE):
        with open(CSV_FILE, mode='w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(DEFAULT_COLUMNS.keys())
    os.makedirs(IMAGE_FOLDER, exist_ok=True)

def add_return(product_name, barcode, return_reason, notes, image_name=''):
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with open(CSV_FILE, mode='a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([timestamp, product_name, barcode, return_reason, notes, image_name])
    st.success(f"已添加退货记录: {product_name} - {barcode}")

@st.cache_data
def load_data():
    try:
        df = pd.read_csv(CSV_FILE)
        df.columns = list(DEFAULT_COLUMNS.keys())[:len(df.columns)]
        return df.rename(columns=DEFAULT_COLUMNS)
    except:
        return pd.DataFrame(columns=DEFAULT_COLUMNS.values())

# ---------------------- 导出 Excel ----------------------
def export_to_excel(df_selected, export_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "退货记录"

    base_headers = list(DEFAULT_COLUMNS.values())[:-1]  # 不包含图片文件名
    max_images = 0

    for _, row in df_selected.iterrows():
        images = [img.strip() for img in str(row['图片文件名']).split(',') if img.strip()]
        max_images = max(max_images, len(images))

    headers = base_headers + [f'图片{i+1}' for i in range(max_images)]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.column_dimensions[get_column_letter(col)].width = 25

    temp_paths = []

    for i, (_, row) in enumerate(df_selected.iterrows(), start=2):
        for j in range(len(base_headers)):
            ws.cell(row=i, column=j+1, value=row[j])

        images = [img.strip() for img in str(row['图片文件名']).split(',') if img.strip()]
        for k, img_name in enumerate(images):
            img_path = os.path.join(IMAGE_FOLDER, img_name)
            if os.path.exists(img_path):
                try:
                    pil_img = PILImage.open(img_path)
                    size = 1000
                    # 缩放图片以适应Excel单元格
                    max_size = (size, size)
                    pil_img.thumbnail(max_size)
                    pil_img = pil_img.resize(max_size, resample=PILImage.LANCZOS)
                    temp_path = f"temp_{i}_{k}.png"
                    pil_img.save(temp_path, format='PNG', dpi=(size, size))
                    w, h = pil_img.size
                    temp_paths.append(temp_path)
                    
                    xl_img = XLImage(temp_path)
                    col_letter = get_column_letter(len(base_headers) + 1 + k)
                    xl_img.anchor = f"{col_letter}{i}"
                    ws.add_image(xl_img)
                    ws.row_dimensions[i].height = 200
                    ws.column_dimensions[col_letter].width = 90 # width 是字符宽度单位

                except:
                    st.warning(f"图片 {img_name} 插入失败")

    try:
        wb.save(export_path)
    finally:
        for path in temp_paths:
            if os.path.exists(path):
                os.remove(path)

# ---------------------- 页面主体 ----------------------
def main():
    st.title("📦 退货记录登记与导出系统")

    init_data()

    # ---------------- 添加退货 ----------------
    st.header("➕ 添加退货记录")

    with st.form("add_form", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            product_name = st.text_input("产品名称*", placeholder="例如: BT-Ice Grey-Twin")
            barcode = st.text_input("条形码*", placeholder="例如: 799392016279")
        with col2:
            return_reason = st.selectbox("退货原因*", RETURN_REASONS)
            notes = st.text_area("备注", placeholder="可填写详细说明")

        uploaded_images = st.file_uploader("上传退货图片（可多选）", type=["jpg", "jpeg", "png"], accept_multiple_files=True)

        submitted = st.form_submit_button("添加记录")
        if submitted:
            if not product_name or not barcode:
                st.error("请填写所有必填字段")
            else:
                image_names = []
                timestamp_prefix = datetime.now().strftime("%Y%m%d_%H%M%S")
                for idx, img in enumerate(uploaded_images):
                    img_name = f"{timestamp_prefix}_{idx}_{img.name}"
                    save_path = os.path.join(IMAGE_FOLDER, img_name)
                    with open(save_path, "wb") as f:
                        f.write(img.getbuffer())
                    image_names.append(img_name)
                image_name_str = ",".join(image_names)
                add_return(product_name, barcode, return_reason, notes, image_name_str)
                st.rerun()

    # ---------------- 展示统计 ----------------
    df = load_data()
    if not df.empty:
        st.header("📊 退货统计概览")

        col1, col2, col3 = st.columns(3)
        with col1:
            st.write("**按产品名称统计**")
            st.dataframe(df["产品名称"].value_counts().reset_index().rename(columns={"index": "产品名称", "产品名称": "退货次数"}))
        with col2:
            st.write("**按条形码统计**")
            st.dataframe(df["条形码"].value_counts().reset_index().rename(columns={"index": "条形码", "条形码": "退货次数"}))
        with col3:
            st.write("**按退货原因统计**")
            st.dataframe(df["退货原因"].value_counts().reset_index().rename(columns={"index": "退货原因", "退货原因": "次数"}))

        st.subheader("📄 所有退货记录")
        st.dataframe(df, use_container_width=True)

        # ---------------- CSV导出 ----------------
        st.subheader("📁 数据导出 - CSV")
        st.download_button(
            label="📥 下载 CSV (中文列名)",
            data=df.to_csv(index=False, encoding='utf-8-sig'),
            file_name="退货记录_中文.csv",
            mime="text/csv"
        )
        raw_df = pd.read_csv(CSV_FILE)
        st.download_button(
            label="📥 下载 CSV (原始列名)",
            data=raw_df.to_csv(index=False),
            file_name="退货记录_原始.csv",
            mime="text/csv"
        )

        # ---------------- Excel导出 ----------------
        st.subheader("📤 数据导出 - Excel（含图片）")

        df_display = df.copy()
        df_display["行号"] = df_display.index.astype(str)
        df_display = df_display[["行号"] + list(df.columns)]
        selected_rows_idx = st.multiselect(
            "选择要导出的行（支持多选）",
            options=df_display["行号"].tolist(),
            format_func=lambda x: f"第 {int(x)+1} 行 - {df_display.iloc[int(x)]['产品名称']}"
        )

        if selected_rows_idx:
            selected_df = df.iloc[[int(i) for i in selected_rows_idx]]
            st.dataframe(selected_df, use_container_width=True)

            if st.button("📤 导出所选记录为 Excel 文件"):
                output_file = "退货导出结果.xlsx"
                export_to_excel(selected_df, output_file)
                with open(output_file, "rb") as f:
                    st.download_button(
                        label="📥 点击下载 Excel 文件",
                        data=f,
                        file_name=output_file,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

if __name__ == "__main__":
    main()
